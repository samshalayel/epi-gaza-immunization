from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, io

app = Flask(__name__)
DATABASE = os.path.join(os.path.dirname(__file__), 'database.db')
EXCEL_SOURCE = r'C:\Users\Administrator\Desktop\androw_new.xlsx'

MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
MONTHS_AR = ['يناير','فبراير','مارس','أبريل','مايو','يونيو','يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']

# Antigens with drop-out rules
ANTIGENS = [
    {'key': 'BCG',    'label': 'BCG',         'color': '#d0e8ff'},
    {'key': 'OPV0',   'label': 'OPV 0',       'color': '#fff0cc'},
    {'key': 'Penta1', 'label': 'Penta 1',     'color': '#d0e8ff'},
    {'key': 'OPV1',   'label': 'OPV 1',       'color': '#fff0cc'},
    {'key': 'PCV1',   'label': 'PCV 1',       'color': '#d0e8ff'},
    {'key': 'Penta2', 'label': 'Penta 2',     'color': '#fff0cc'},
    {'key': 'OPV2',   'label': 'OPV 2',       'color': '#d0e8ff'},
    {'key': 'PCV2',   'label': 'PCV 2',       'color': '#fff0cc'},
    {'key': 'Penta3', 'label': 'Penta 3',     'color': '#d0e8ff'},
    {'key': 'OPV3',   'label': 'OPV 3',       'color': '#fff0cc'},
    {'key': 'PCV3',   'label': 'PCV 3',       'color': '#d0e8ff'},
    {'key': 'IPV',    'label': 'IPV',         'color': '#fff0cc'},
    {'key': 'MR1',    'label': 'MR 1 (MMR1)', 'color': '#d0e8ff'},
    {'key': 'MR2',    'label': 'MR 2 (MMR2)', 'color': '#fff0cc'},
    {'key': 'VitA1',  'label': 'Vit A (1st)', 'color': '#d0e8ff'},
    {'key': 'VitA2',  'label': 'Vit A (2nd)', 'color': '#fff0cc'},
]

# Drop-out definitions: (label, numerator_key, denominator_key)
DROPOUTS = [
    ('Dropout 1: Penta1→Penta3',  'Penta1', 'Penta3'),
    ('Dropout 2: BCG→MR1',        'BCG',    'MR1'),
    ('Dropout 3: MR1→MR2',        'MR1',    'MR2'),
]

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS facilities (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sno INTEGER, governorate TEXT, name TEXT,
        type TEXT, provider TEXT, status TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS target_population (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        facility_id INTEGER, year INTEGER,
        catchment_pop INTEGER DEFAULT 0,
        si_percent REAL DEFAULT 3.2,
        UNIQUE(facility_id, year)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS immunization_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        facility_id INTEGER, year INTEGER,
        antigen TEXT, month INTEGER,
        doses INTEGER DEFAULT 0,
        UNIQUE(facility_id, year, antigen, month)
    )''')
    conn.commit()
    # Import facilities if table is empty
    c.execute('SELECT COUNT(*) FROM facilities')
    if c.fetchone()[0] == 0:
        import_facilities(conn)
    conn.close()

def import_facilities(conn=None):
    close = False
    if conn is None:
        conn = get_db()
        close = True
    try:
        df = pd.read_excel(EXCEL_SOURCE, sheet_name='Sheet1', header=1)
        c = conn.cursor()
        for _, row in df.iterrows():
            if pd.notna(row.get('Name of the Facility')):
                c.execute('''INSERT OR IGNORE INTO facilities
                    (sno, governorate, name, type, provider, status) VALUES (?,?,?,?,?,?)''',
                    (row.get('S.No.'), str(row.get('Governorate','')),
                     str(row.get('Name of the Facility','')),
                     str(row.get('Type of Facility','')),
                     str(row.get('Provider','')),
                     str(row.get('Status of Functionality',''))))
        conn.commit()
    except Exception as e:
        print(f"Import error: {e}")
    if close:
        conn.close()

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    year = request.args.get('year', 2025, type=int)
    search = request.args.get('search', '')
    gov = request.args.get('governorate', '')
    conn = get_db()
    query = 'SELECT * FROM facilities WHERE 1=1'
    params = []
    if search:
        query += ' AND name LIKE ?'
        params.append(f'%{search}%')
    if gov:
        query += ' AND governorate = ?'
        params.append(gov)
    query += ' ORDER BY governorate, sno'
    facilities = conn.execute(query, params).fetchall()
    govs = conn.execute('SELECT DISTINCT governorate FROM facilities ORDER BY governorate').fetchall()
    conn.close()
    return render_template('index.html', facilities=facilities, year=year,
                           governorates=govs, search=search, sel_gov=gov)

@app.route('/facility/<int:fid>', methods=['GET','POST'])
def facility(fid):
    year = request.args.get('year', 2025, type=int)
    conn = get_db()
    fac = conn.execute('SELECT * FROM facilities WHERE id=?', (fid,)).fetchone()
    if not fac:
        conn.close()
        return redirect(url_for('index'))

    if request.method == 'POST':
        year = int(request.form.get('year', 2025))
        catchment = int(request.form.get('catchment_pop', 0) or 0)
        si = float(request.form.get('si_percent', 3.2) or 3.2)
        conn.execute('''INSERT INTO target_population (facility_id, year, catchment_pop, si_percent)
            VALUES (?,?,?,?) ON CONFLICT(facility_id,year) DO UPDATE SET
            catchment_pop=excluded.catchment_pop, si_percent=excluded.si_percent''',
            (fid, year, catchment, si))
        for ag in ANTIGENS:
            for m in range(1, 13):
                val = int(request.form.get(f"{ag['key']}_{m}", 0) or 0)
                conn.execute('''INSERT INTO immunization_data (facility_id,year,antigen,month,doses)
                    VALUES (?,?,?,?,?) ON CONFLICT(facility_id,year,antigen,month) DO UPDATE SET doses=excluded.doses''',
                    (fid, year, ag['key'], m, val))
        conn.commit()
        conn.close()
        return redirect(url_for('facility', fid=fid, year=year, saved=1))

    # Load existing data
    tp = conn.execute('SELECT * FROM target_population WHERE facility_id=? AND year=?',
                      (fid, year)).fetchone()
    rows = conn.execute('SELECT antigen, month, doses FROM immunization_data WHERE facility_id=? AND year=?',
                        (fid, year)).fetchall()
    conn.close()

    data = {}
    for r in rows:
        data[(r['antigen'], r['month'])] = r['doses']

    saved = request.args.get('saved', 0)
    return render_template('facility.html', fac=fac, year=year, tp=tp,
                           antigens=ANTIGENS, months=MONTHS, data=data,
                           dropouts=DROPOUTS, saved=saved)

@app.route('/add_facility', methods=['GET','POST'])
def add_facility():
    if request.method == 'POST':
        conn = get_db()
        conn.execute('''INSERT INTO facilities (governorate, name, type, provider, status)
            VALUES (?,?,?,?,?)''',
            (request.form['governorate'], request.form['name'],
             request.form['type'], request.form['provider'], 'Functional'))
        conn.commit()
        conn.close()
        return redirect(url_for('index'))
    return render_template('add_facility.html')

@app.route('/reimport')
def reimport():
    import_facilities()
    return redirect(url_for('index'))

@app.route('/export/<int:fid>')
def export_facility(fid):
    year = request.args.get('year', 2025, type=int)
    conn = get_db()
    fac = conn.execute('SELECT * FROM facilities WHERE id=?', (fid,)).fetchone()
    tp = conn.execute('SELECT * FROM target_population WHERE facility_id=? AND year=?',
                      (fid, year)).fetchone()
    rows = conn.execute('SELECT antigen, month, doses FROM immunization_data WHERE facility_id=? AND year=?',
                        (fid, year)).fetchall()
    conn.close()

    data = {}
    for r in rows:
        data[(r['antigen'], r['month'])] = r['doses']

    wb = build_excel(fac, year, tp, data)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = str(fac['name']).replace(' ', '_').replace('/', '-')
    return send_file(buf, as_attachment=True,
                     download_name=f"EPI_{safe_name}_{year}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export_all')
def export_all():
    year = request.args.get('year', 2025, type=int)
    conn = get_db()
    facilities = conn.execute('SELECT * FROM facilities ORDER BY governorate, sno').fetchall()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for fac in facilities:
        tp = conn.execute('SELECT * FROM target_population WHERE facility_id=? AND year=?',
                          (fac['id'], year)).fetchone()
        rows = conn.execute('SELECT antigen,month,doses FROM immunization_data WHERE facility_id=? AND year=?',
                            (fac['id'], year)).fetchall()
        data = {(r['antigen'], r['month']): r['doses'] for r in rows}
        sheet_name = str(fac['name'])[:31].replace('/', '-').replace('\\','-').replace('*','').replace('[','').replace(']','').replace(':','').replace('?','')
        add_sheet_to_wb(wb, sheet_name, fac, year, tp, data)

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"EPI_All_Facilities_{year}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── Excel Builder ────────────────────────────────────────────────────────────

HDR_FILL  = PatternFill('solid', fgColor='1FA2C8')
HDR2_FILL = PatternFill('solid', fgColor='3ABCD8')
BLUE_FILL = PatternFill('solid', fgColor='D0E8FF')
YELL_FILL = PatternFill('solid', fgColor='FFF0CC')
GRN_FILL  = PatternFill('solid', fgColor='C6EFCE')
RED_FILL  = PatternFill('solid', fgColor='FFC7CE')
WHT_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
BLD_FONT  = Font(name='Arial', bold=True, size=9)
NRM_FONT  = Font(name='Arial', size=9)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='888888')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(ws, r, c, value='', fill=None, font=None, align=CTR, border=BORDER, number_format=None):
    cell = ws.cell(row=r, column=c, value=value)
    if fill:   cell.fill = fill
    if font:   cell.font = font
    if align:  cell.alignment = align
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell

def build_excel(fac, year, tp, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(fac['name'])[:31]
    add_sheet_content(ws, fac, year, tp, data)
    return wb

def add_sheet_to_wb(wb, sheet_name, fac, year, tp, data):
    ws = wb.create_sheet(title=sheet_name)
    add_sheet_content(ws, fac, year, tp, data)

def add_sheet_content(ws, fac, year, tp, data):
    # ── Title ──
    ws.merge_cells('A1:AO1')
    t = ws['A1']
    t.value = f"Routine Immunization Monthly Monitoring  |  {fac['name']}  |  Year: {year}"
    t.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    t.fill = HDR_FILL
    t.alignment = CTR

    # ── Facility info row ──
    ws.merge_cells('A2:D2')
    ws['A2'].value = f"Governorate: {fac['governorate']}"
    ws['A2'].font = BLD_FONT
    ws['A2'].fill = HDR2_FILL
    ws['A2'].font = WHT_FONT
    ws['A2'].alignment = CTR
    ws.merge_cells('E2:H2')
    ws['E2'].value = f"Provider: {fac['provider']}"
    ws['E2'].font = WHT_FONT
    ws['E2'].fill = HDR2_FILL
    ws['E2'].alignment = CTR
    ws.merge_cells('I2:L2')
    ws['I2'].value = f"Type: {fac['type']}"
    ws['I2'].font = WHT_FONT
    ws['I2'].fill = HDR2_FILL
    ws['I2'].alignment = CTR

    # ── Target Population Row ──
    catchment = tp['catchment_pop'] if tp else 0
    si        = tp['si_percent']    if tp else 3.2
    target_u1 = int(catchment * si / 100)

    ws.merge_cells('A3:C3')
    ws['A3'].value = 'Catchment Population'
    ws['A3'].font = BLD_FONT; ws['A3'].fill = HDR2_FILL
    ws['A3'].font = WHT_FONT; ws['A3'].alignment = CTR

    ws['D3'].value = catchment
    ws['D3'].font = BLD_FONT; ws['D3'].alignment = CTR; ws['D3'].border = BORDER
    ws['D3'].number_format = '#,##0'

    ws.merge_cells('E3:G3')
    ws['E3'].value = '% Surviving Infants (SI)'
    ws['E3'].font = WHT_FONT; ws['E3'].fill = HDR2_FILL; ws['E3'].alignment = CTR

    ws['H3'].value = si / 100
    ws['H3'].number_format = '0.0%'
    ws['H3'].font = BLD_FONT; ws['H3'].alignment = CTR; ws['H3'].border = BORDER

    ws.merge_cells('I3:K3')
    ws['I3'].value = 'Target (Under 1 yr)'
    ws['I3'].font = WHT_FONT; ws['I3'].fill = HDR2_FILL; ws['I3'].alignment = CTR

    ws['L3'].value = f'=D3*H3'
    ws['L3'].number_format = '#,##0'
    ws['L3'].font = BLD_FONT; ws['L3'].alignment = CTR; ws['L3'].border = BORDER

    # ── Column Headers ──
    # Row 4: Antigen | Jan (Tot, Cum) | Feb (Tot, Cum) ... | Dec (Tot,Cum) | Annual | Coverage%
    ROW_HDR = 4
    ws.merge_cells(f'A{ROW_HDR}:A{ROW_HDR+1}')
    cell_style(ws, ROW_HDR, 1, 'Antigen / Vaccine', HDR_FILL, WHT_FONT)

    COL_START = 2
    month_col = {}  # month -> (tot_col, cum_col)
    for i, m in enumerate(MONTHS):
        tc = COL_START + i * 2
        cc = tc + 1
        month_col[i+1] = (tc, cc)
        ws.merge_cells(f'{get_column_letter(tc)}{ROW_HDR}:{get_column_letter(cc)}{ROW_HDR}')
        cell_style(ws, ROW_HDR, tc, m, HDR_FILL, WHT_FONT)
        cell_style(ws, ROW_HDR+1, tc, 'Tot.', HDR2_FILL, WHT_FONT)
        cell_style(ws, ROW_HDR+1, cc, 'Cum.', HDR2_FILL, WHT_FONT)

    ANN_COL = COL_START + 24      # Annual total
    COV_COL = ANN_COL + 1         # Coverage %
    DO_COL  = COV_COL + 1         # Dropout %

    ws.merge_cells(f'{get_column_letter(ANN_COL)}{ROW_HDR}:{get_column_letter(ANN_COL)}{ROW_HDR+1}')
    cell_style(ws, ROW_HDR, ANN_COL, 'Annual Total', HDR_FILL, WHT_FONT)
    ws.merge_cells(f'{get_column_letter(COV_COL)}{ROW_HDR}:{get_column_letter(COV_COL)}{ROW_HDR+1}')
    cell_style(ws, ROW_HDR, COV_COL, 'Coverage %', HDR_FILL, WHT_FONT)

    # ── Data Rows ──
    DATA_START = ROW_HDR + 2
    antigen_rows = {}   # key -> excel row

    for ag_i, ag in enumerate(ANTIGENS):
        r = DATA_START + ag_i
        antigen_rows[ag['key']] = r
        fill = BLUE_FILL if ag_i % 2 == 0 else YELL_FILL

        cell_style(ws, r, 1, ag['label'], fill, BLD_FONT, Alignment(horizontal='left', vertical='center'))

        for m in range(1, 13):
            tc, cc = month_col[m]
            val = data.get((ag['key'], m), 0)
            cell_style(ws, r, tc, val, fill, NRM_FONT, number_format='#,##0')

            # Cumulative formula
            if m == 1:
                cum_formula = f'={get_column_letter(tc)}{r}'
            else:
                prev_cc = get_column_letter(month_col[m-1][1])
                cum_formula = f'={prev_cc}{r}+{get_column_letter(tc)}{r}'
            cell_style(ws, r, cc, cum_formula, fill, Font(name='Arial', size=9, italic=True),
                       number_format='#,##0')

        # Annual = sum of all Tot columns
        tot_cols = '+'.join(get_column_letter(month_col[m][0])+str(r) for m in range(1,13))
        cell_style(ws, r, ANN_COL, f'={tot_cols}', fill, BLD_FONT, number_format='#,##0')

        # Coverage % = Annual / Target
        cell_style(ws, r, COV_COL, f'=IF($L$3>0,{get_column_letter(ANN_COL)}{r}/$L$3,0)',
                   fill, Font(name='Arial', size=9, color='000000'), number_format='0.0%')

    # ── Drop-out Rows ──
    DO_START = DATA_START + len(ANTIGENS) + 1
    do_header_row = DO_START - 1
    ws.merge_cells(f'A{do_header_row}:{get_column_letter(DO_COL)}{do_header_row}')
    cell_style(ws, do_header_row, 1, 'Drop-out Rates', HDR_FILL, WHT_FONT,
               Alignment(horizontal='left', vertical='center'))

    for di, (label, num_key, den_key) in enumerate(DROPOUTS):
        r = DO_START + di
        fill = BLUE_FILL if di % 2 == 0 else YELL_FILL
        cell_style(ws, r, 1, label, fill, BLD_FONT,
                   Alignment(horizontal='left', vertical='center'))

        num_r = antigen_rows.get(num_key)
        den_r = antigen_rows.get(den_key)
        ann_num = f'{get_column_letter(ANN_COL)}{num_r}'
        ann_den = f'{get_column_letter(ANN_COL)}{den_r}'

        # Monthly dropout
        for m in range(1, 13):
            tc, cc = month_col[m]
            num_cc = get_column_letter(month_col[m][1])
            den_cc = get_column_letter(month_col[m][1])
            num_cell = f'{get_column_letter(month_col[m][1])}{num_r}'
            den_cell = f'{get_column_letter(month_col[m][1])}{den_r}'
            formula_do = f'=IF({num_cell}>0,({num_cell}-{den_cell})/{num_cell},0)'
            do_cell = ws.cell(row=r, column=tc, value=formula_do)
            do_cell.number_format = '0.0%'
            do_cell.fill = fill
            do_cell.font = NRM_FONT
            do_cell.border = BORDER
            ws.merge_cells(f'{get_column_letter(tc)}{r}:{get_column_letter(cc)}{r}')

        # Annual dropout
        do_ann = f'=IF({ann_num}>0,({ann_num}-{ann_den})/{ann_num},0)'
        ann_cell = ws.cell(row=r, column=ANN_COL, value=do_ann)
        ann_cell.number_format = '0.0%'
        ann_cell.fill = fill; ann_cell.font = BLD_FONT; ann_cell.border = BORDER

        # Conditional note
        note_cell = ws.cell(row=r, column=COV_COL,
                            value=f'=IF({get_column_letter(ANN_COL)}{r}>0.1,"HIGH ⚠","OK ✓")')
        note_cell.fill = fill; note_cell.font = BLD_FONT
        note_cell.alignment = CTR; note_cell.border = BORDER

    # ── Achievement Legend ──
    LEG_ROW = DO_START + len(DROPOUTS) + 2
    ws.merge_cells(f'A{LEG_ROW}:C{LEG_ROW}')
    cell_style(ws, LEG_ROW, 1, 'Coverage Achievement Legend:', None, BLD_FONT,
               Alignment(horizontal='left'))
    thresholds = [('≥100%', '00B050'), ('≥75%','92D050'), ('≥50%','FFEB84'), ('≥45%','FF9900'), ('<45%','FF0000')]
    for ti, (lbl, color) in enumerate(thresholds):
        c = 4 + ti
        cell = ws.cell(row=LEG_ROW, column=c, value=lbl)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(name='Arial', size=9, bold=True)
        cell.alignment = CTR
        cell.border = BORDER

    # ── Formula Notes ──
    NOTE_ROW = LEG_ROW + 2
    ws.merge_cells(f'A{NOTE_ROW}:H{NOTE_ROW}')
    ws[f'A{NOTE_ROW}'].value = 'Formulas:  Drop-out # = First Dose − Last Dose  |  Drop-out % = Drop-out # ÷ First Dose × 100  |  Coverage % = Annual Doses ÷ Target Population × 100'
    ws[f'A{NOTE_ROW}'].font = Font(name='Arial', size=8, italic=True, color='444444')

    # ── Column widths ──
    ws.column_dimensions['A'].width = 22
    for m in range(1, 13):
        tc, cc = month_col[m]
        ws.column_dimensions[get_column_letter(tc)].width = 7
        ws.column_dimensions[get_column_letter(cc)].width = 7
    ws.column_dimensions[get_column_letter(ANN_COL)].width = 12
    ws.column_dimensions[get_column_letter(COV_COL)].width = 12
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = f'B{DATA_START}'

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
